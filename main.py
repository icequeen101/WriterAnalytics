# This is a sample Python script. Testing commit

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import docx2txt


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    # print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

    # read in word file
    result = docx2txt.process("Hello_World.docx")
    result2 = docx2txt.process("C:\\Users\\lwu\\Hello World.docx")
    print('hi')
    print(result)
    print(result2)

    # Cleaning text and lower casing all words
    for char in ',.!?-–\n':  # TO DO: ;;""''
        result = result.replace(char, ' ')
    result = result.lower()

    # split returns a list of words delimited by sequences of whitespace (including tabs, newlines, etc, like re's \s)
    word_list = result.split()
    print(word_list)

    from collections import Counter
    Counter(word_list).most_common()
    print(Counter(word_list).most_common())
    c = Counter(word_list)
    # c.total()

    # Sum of all
    print(sum(c.values()))
    # print(Counter(word_list).total())

    import xlsxwriter
    workbook = xlsxwriter.Workbook('Writer Analytics Sheet.xlsx')
    worksheet = workbook.add_worksheet("My Sheet")
    # Some data we want to write to the worksheet.
    scores = (
        ['ankit', 1000],
        ['rahul', 100],
        ['priya', 300],
        ['harshita', 50],
    )

    # Start from the first cell. Rows and
    # columns are zero indexed.
    row = 0
    col = 0

    # Iterate over the data and write it out row by row.
    for name, score in (scores):
        worksheet.write(row, col, name)
        worksheet.write(row, col + 1, score)
        row += 1

    workbook.close()

    import openpyxl

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter


    wb = Workbook()

    dest_filename = 'empty_book.xlsx'

    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
        ...
        ws1.append(range(600))

    ws2 = wb.create_sheet(title="Pi")

    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        ...
        for col in range(27, 54):
            ...
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)

    wb.save(filename=dest_filename)

from openpyxl import load_workbook
from openpyxl import load_workbook

wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
# max = ws.max_row
# for row, entry in enumerate(data1, start=1):
#     st.cell(row=row + max, column=1, value=entry)

from openpyxl import load_workbook

workbook_name = 'Example3.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

# New data to write:
new_companies = [['name3','address3'], ['name4','address4','tel4','web4']]

for info in new_companies:
    page.append(info)

wb.save(filename=workbook_name)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
